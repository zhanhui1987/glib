#!/usr/bin/env python3
# coding: utf-8

# @Time    : 2018/5/6 1:09
# @Author  : zhanhui
# @File    : my_excel_x.py


"""
    利用python的openpyxl模块，处理xlsx后缀的excel文件。
"""


import re
import openpyxl
import os

from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils.exceptions import IllegalCharacterError

# 边框尺寸设置
BORDER = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))
# 定义字体
FONT = {
    "name": "Times New Roman",  # 字体
    "size": 12,  # 文字大小
    "bold": True,  # 是否加粗
}

# 图案填充
PATTERN_FILL = {
    "fill_type": "lightUp",
    "start_color": "00339966",
    "end_color": "00339966",
}

# 每个sheet页面的表头中，每个表格的宽度
SHEET_HEADER_WIDTH_DICT = {
    "sheet1": {
        "title1": 15.0,
        "title2": 27.0,
    }
}


class ExcelRead(object):
    # 初始化excel对象
    def __init__(self, file_name):
        real_file_name = os.path.realpath(file_name)
        if real_file_name.endswith(r".xlsx"):
            self._file_name = real_file_name

        self._wb = None

    # 读取excel中的内容
    def _read_excel(self):
        try:
            self._wb = openpyxl.load_workbook(self._file_name)
        except Exception as e:
            print(e)

    # 获取excel的sheet列表
    def _get_sheets(self):
        self._read_excel()
        if self._wb is not None and isinstance(self._wb, openpyxl.workbook.workbook.Workbook):
            self._sheet_names = self._wb.sheetnames
        else:
            self._sheet_names = []

    # 返回excel的sheet列表
    @property
    def sheet_names(self):
        self._get_sheets()
        return self._sheet_names

    # 获取sheet中的内容
    def read(self, has_header=True, read_sheet_name=None):
        self._get_sheets()

        # 定义excel中数据保存的空数组
        excel_data_list = []
        if self._wb is not None and isinstance(self.sheet_names, list) and self.sheet_names:
            for sheet_name in self.sheet_names:
                # 若有传入的去读的read_sheet，则将不是该read_sheet的sheet页面跳过。
                if read_sheet_name is not None and sheet_name != read_sheet_name:
                    continue

                # 定义每一sheet读取的数据结构
                sheet_data_dict = {
                    "sheet_name": sheet_name,
                    "sheet_data": [],
                    "sheet_header": [],
                }

                # 获取sheet_name对应的sheet对象的所有行
                sheet_rows = self._wb[sheet_name].rows

                if sheet_rows:
                    # 定义表头数组，用于获取表头信息
                    header_list = []

                    # 循环所有行，取出每一行的内容；需要判断是否有表头，有的话，从中取出第一行作为表头
                    for row_index, row in enumerate(sheet_rows):
                        # 若excel有表头，且当前行数是第一行，则取出第一行每一个cell的值，作为表头的内容。
                        if has_header and row_index == 0:
                            for cell in row:
                                # 获取每个cell的值，并对其进行strip处理
                                header_list.append(_get_cell_value(cell))
                        else:
                            # 若当前是第一行，且excel文件没有表头，则将row中数据的长度，来作为表头内容
                            if row_index == 0 and not header_list:
                                header_list = [i for i in range(len(row))]

                            # 定义每一行数据的字典
                            row_data = dict()

                            # 处理每一行的cell，将其内容与表头对应起来
                            for column, cell in enumerate(row):
                                row_data[header_list[column]] = _get_cell_value(cell)

                            sheet_data_dict["sheet_data"].append(row_data)

                    # 将表头信息进行保存
                    sheet_data_dict["sheet_header"] = header_list

                # 将获取到的sheet页面的数据，保存到excel数据列表中
                excel_data_list.append(sheet_data_dict)

        return excel_data_list


class ExcelWrite(object):
    # 初始化excel对象
    def __init__(self, file_name):
        # 获取excel文件的全路径，且应保证文件是xlsx文件。
        real_file_name = os.path.realpath(file_name)
        if real_file_name.endswith(r".xlsx"):
            self._file_name = real_file_name

    def write(self, data, save_header=True):
        # save_header表明是否需要将表头列表保存到excel文件中，total_header_style_dict用来定义表头的格式。

        # 确保文件名存在
        if self._file_name is None:
            raise FileNotFoundError("文件名不存在！")

        # 初始化data中表头的格式，资产核查结果excel文件默认都有这些格式。
        self._set_data_header_style(data)

        # 初始化excel对象
        wb = openpyxl.Workbook()

        # 定义excel中非法字符的匹配方法
        illegal_characters_re = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

        for i, sheet_row in enumerate(data):
            # 获取sheet_name、sheet_data、sheet_header等信息
            sheet_name = sheet_row.get("sheet_name")
            sheet_data = sheet_row.get("sheet_data")
            sheet_header = sheet_row.get("sheet_header")

            # --- 数据验证，并将不存在的值设置为默认值 -- #
            # 确保sheet_data是数组结构，且不为空
            if not isinstance(sheet_data, list) and sheet_data:
                continue

            # 若没有设置sheet_name，则给其赋一个默认值
            if sheet_name is None:
                sheet_name = "Sheet%s" % i

            # 若没有设置sheet_header，则给其赋一个默认值。默认是sheet_data的长度。
            if not isinstance(sheet_header, list):
                sheet_header = [i for i in range(len(sheet_data))]

            # --- 创建新的sheet页面 --- #
            # 新建sheet，若是第一个sheet，则使用默认创建的sheet，将其sheet_name进行修改；不是第一个的话就新建sheet。
            if i:
                new_sheet = wb.create_sheet(title=sheet_name, index=i)
            else:
                default_sheet_name = wb.sheetnames[0]
                new_sheet = wb[default_sheet_name]
                new_sheet.title = sheet_name

            # --- 将表头写入sheet中 --- #
            # 若save_header为True且表头数据不为空，则需要将sheet_header写入到sheet中；否则不需要。
            if save_header and sheet_header:
                new_sheet.append(sheet_header)

                # 获取sheet_header_style，若其存在，则需要将表头的格式进行修改
                header_style_dict = sheet_row.get("header_style_dict")
                if isinstance(header_style_dict, dict):
                    # 对excel文件的表头的格式进行处理。
                    self._set_excel_header_style(new_sheet, header_style_dict)

            # --- 将sheet_data中的内容写入sheet中 --- #

            if sheet_data:
                for row in sheet_data:
                    save_data_list = []

                    # 循环sheet_header，获取每一行数据对应表头的值的列表。
                    for header in sheet_header:
                        value = row.get(header, "")

                        # 将value转换成str类型
                        save_data_list.append(str(value))

                    # 将每一行数据保存到sheet中
                    try:
                        new_sheet.append(save_data_list)
                    except IllegalCharacterError:
                        # 若捕获到IllegalCharacterError，则需要将save_data_list中的值去除非法字符
                        new_save_data_list = []
                        for data in save_data_list:
                            new_data = illegal_characters_re.sub(r"", data)
                            new_save_data_list.append(new_data)
                        new_sheet.append(new_save_data_list)

        # 将上述创建的sheet和设置的数据进行保存
        wb.save(self._file_name)

    @staticmethod
    def _set_data_header_style(data):
        # 设置data中，每个sheet对应的表头的格式

        # 需要将sheet_name的数字后缀去除
        int_tail_re = re.compile('\d+$')

        for row in data:
            header_style_dict = dict()

            # 设置字体格式
            if FONT is not None:
                header_style_dict["font"] = FONT

            # 设置图案填充颜色
            if PATTERN_FILL is not None:
                header_style_dict["pattern_fill"] = PATTERN_FILL

            # 设置宽度格式，根据sheet_name获取其对应的表头宽度
            if isinstance(SHEET_HEADER_WIDTH_DICT, dict):
                sheet_name = row.get("sheet_name")
                if sheet_name is not None:
                    sheet_name = int_tail_re.sub("", sheet_name)
                    header_style_dict["header_width_dict"] = SHEET_HEADER_WIDTH_DICT.get(sheet_name)

            # 设置边框格式
            if BORDER is not None:
                header_style_dict["border"] = BORDER

            row["header_style_dict"] = header_style_dict

    @staticmethod
    def _set_excel_header_style(new_sheet, header_style_dict):
        # 若传入了表头的格式，则需要对其进行处理

        if isinstance(header_style_dict, dict):
            font = header_style_dict.get("font")
            pattern_fill = header_style_dict.get("pattern_fill")
            header_width_dict = header_style_dict.get("header_width_dict")
            border = header_style_dict.get("border")

            # --- 修改header的样式 --- #
            # 获取字体设置
            header_font = None
            if isinstance(font, dict) and font:
                header_font = Font(**font)

            # 获取图案填充设置
            header_pattern_fill = None
            if isinstance(pattern_fill, dict) and pattern_fill:
                header_pattern_fill = PatternFill(**pattern_fill)

            # 获取new_sheet所有的列
            columns = new_sheet.columns
            # 循环每一列
            for col in columns:
                try:
                    # 每一列的第一行，是该列的表头cell
                    header_cell = col[0]

                    # --- 设置表头cell的宽度 --- #
                    # 获取表头cell的名称，并通过该名称获取该表头对应的宽度，宽度存在的话则对表头的宽度进行调整
                    if isinstance(header_width_dict, dict):
                        header_cell_width = header_width_dict.get(_get_cell_value(header_cell))
                        if header_cell_width is not None and isinstance(header_cell_width, float):
                            # 设置表头cell所在列的宽度
                            new_sheet.column_dimensions[header_cell.column].width = header_cell_width

                    # --- 设置表头cell的字体 --- #
                    if header_font is not None:
                        header_cell.font = header_font

                    # --- 设置表头cell的图案填充 --- #
                    if header_pattern_fill is not None:
                        header_cell.fill = header_pattern_fill

                    # --- 设置表头cell的边界 --- #
                        header_cell.border = border

                except Exception as e:
                    print(e)


def _get_cell_value(cell):
    # 获取cell中的值，并对其进行strip处理

    try:
        cell_value = cell.value.strip()
    except AttributeError:
        cell_value = None

    return cell_value


def main():
    # 以核查结果excel文件为例

    file_name = r"F:\1.xlsx"

    er = ExcelRead(file_name)
#    print(er.sheet_names)

    # 读取excel中所有的数据
    data = er.read()
#    print(data)

    # 将读取出的excel文件的数据，存到另外一个excel文件中。
    file_name2 = r"F:\2.xlsx"
    ew = ExcelWrite(file_name2)

    ew.write(data)


if __name__ == "__main__":
    main()
